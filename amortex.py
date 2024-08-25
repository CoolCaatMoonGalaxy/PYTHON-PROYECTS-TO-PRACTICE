import csv
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
import numpy as np

# Función para graficar barras agrupadas
def grafico_barras_agrupadas(detalles_pagos):
    elementos = ["Cuota", "Saldo", "Interes", "Amortizacion"]
    
    pagos = [pago[0] for pago in detalles_pagos]
    cuotas = [pago[1] for pago in detalles_pagos]
    saldos = [pago[2] for pago in detalles_pagos]
    intereses = [pago[3] for pago in detalles_pagos]
    amortizaciones = [pago[4] for pago in detalles_pagos]
    
    fig, ax = plt.subplots(layout='constrained')
    width = 0.2
    x = np.arange(len(pagos)) 

    ax.bar(x, cuotas, width, label="Cuota", color='blue')
    ax.bar(x + width, saldos, width, label="Saldo", color='green')
    ax.bar(x + 2 * width, intereses, width, label="Interes", color='orange')
    ax.bar(x + 3 * width, amortizaciones, width, label="Amortizacion", color='red')

    ax.set_ylabel('Cantidad')
    ax.set_title('Evolución de la Amortización')
    ax.set_xticks(x + 1.5 * width, pagos)
    ax.legend(loc='upper left', ncols=4)
    plt.show()

# Gráfico Broken Barh
def grafico_broken_barh(detalles_pagos):
    fig, ax = plt.subplots(layout='constrained')
    for i, pago in enumerate(detalles_pagos):
        cuota, saldo, interes, amortizacion = pago[1:]
        ax.broken_barh([(i, cuota)], (0, 10), facecolors='blue')
        ax.broken_barh([(i, saldo)], (10, 20), facecolors='green')
        ax.broken_barh([(i, interes)], (20, 30), facecolors='orange')
        ax.broken_barh([(i, amortizacion)], (30, 40), facecolors='red')
    ax.set_yticks([5, 15, 25, 35], ["Cuota", "Saldo", "Interes", "Amortizacion"])
    ax.set_xlabel('Pago')
    ax.set_title('Evolución de la Amortización')
    plt.show()

# Gráfico de Coherencia de dos señales
def grafico_coherencia(detalles_pagos):
    cuotas = [pago[1] for pago in detalles_pagos]
    intereses = [pago[3] for pago in detalles_pagos]
    
    fig, ax = plt.subplots(layout='constrained')
    ax.cohere(cuotas, intereses, NFFT=256, Fs=1)
    ax.set_title('Coherencia entre Cuotas e Intereses')
    plt.show()

# Gráfico con fill_between
def grafico_fill_between(detalles_pagos):
    pagos = [pago[0] for pago in detalles_pagos]
    saldos = [pago[2] for pago in detalles_pagos]

    fig, ax = plt.subplots(layout='constrained')
    ax.plot(pagos, saldos, color='green')
    ax.fill_between(pagos, 0, saldos, where=(np.array(saldos) > 0), color='green', alpha=0.3)
    ax.set_xlabel('Pago')
    ax.set_ylabel('Saldo')
    ax.set_title('Saldo Restante en Cada Pago')
    plt.show()

# Gráfico de Densidad Espectral Cruzada (CSD)
def grafico_csd(detalles_pagos):
    cuotas = [pago[1] for pago in detalles_pagos]
    amortizaciones = [pago[4] for pago in detalles_pagos]

    fig, ax = plt.subplots(layout='constrained')
    ax.csd(cuotas, amortizaciones, NFFT=256, Fs=1)
    ax.set_title('Densidad Espectral Cruzada entre Cuotas y Amortizaciones')
    plt.show()

# Gráfico de barras agrupadas
def grafico_barras_agrupadas(detalles_pagos):
    elementos = ["Cuota", "Saldo", "Interes", "Amortizacion"]
    
    pagos = [pago[0] for pago in detalles_pagos]
    cuotas = [pago[1] for pago in detalles_pagos]
    saldos = [pago[2] for pago in detalles_pagos]
    intereses = [pago[3] for pago in detalles_pagos]
    amortizaciones = [pago[4] for pago in detalles_pagos]
    
    fig, ax = plt.subplots(layout='constrained')
    width = 0.2
    x = np.arange(len(pagos)) 

    ax.bar(x, cuotas, width, label="Cuota", color='blue')
    ax.bar(x + width, saldos, width, label="Saldo", color='green')
    ax.bar(x + 2 * width, intereses, width, label="Interes", color='orange')
    ax.bar(x + 3 * width, amortizaciones, width, label="Amortizacion", color='red')

    ax.set_ylabel('Cantidad')
    ax.set_title('Evolución de la Amortización')
    ax.set_xticks(x + 1.5 * width, pagos)
    ax.legend(loc='upper left', ncols=4)
    plt.show()
    
# Función principal para seleccionar el gráfico
def seleccionar_grafico(tipo, detalles_pagos):
    if tipo == '1':
        grafico_barras_agrupadas(detalles_pagos)
    elif tipo == '2':
        grafico_broken_barh(detalles_pagos)
    elif tipo == '3':
        grafico_coherencia(detalles_pagos)
    elif tipo == '4':
        grafico_fill_between(detalles_pagos)
    elif tipo == '5':
        grafico_csd(detalles_pagos)
    else:
        print("Tipo de gráfico no reconocido. Por favor, elija un tipo válido.")

# Función para crear lista de datos para exportación
def crear_lista_datos(detalles_pagos):
    lista_datos = [["Pago", "Cuota", "Saldo", "Interes", "Amortizacion"]]
    for pago in detalles_pagos:
        lista_datos.append(list(pago))
    return lista_datos

# Función para exportar a PDF
def exportar_a_pdf(detalles_pagos, file_path, color1=colors.beige, color2=colors.white):
    try:
        doc = SimpleDocTemplate(file_path, pagesize=letter)
        elements = []

        data = [["Pago", "Cuota", "Saldo", "Interes", "Amortizacion"]]
        for i, (pago, cuota, saldo, interes, amortizacion) in enumerate(detalles_pagos):
            data.append([Paragraph(str(pago), ParagraphStyle(name='Normal')), 
                        Paragraph(str(cuota), ParagraphStyle(name='Normal')), 
                        Paragraph(str(saldo), ParagraphStyle(name='Normal')), 
                        Paragraph(str(interes), ParagraphStyle(name='Normal')), 
                        Paragraph(str(amortizacion), ParagraphStyle(name='Normal'))])
        tabla = Table(data)
        tabla.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                ('BACKGROUND', (0, 1), (-1, -1), color1),
                                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color1, color2]),
                                ('GRID', (0, 0), (-1, -1), 1, colors.magenta)]))
        elements.append(tabla)
        doc.build(elements)
        print(f"Los detalles de cada pago se han exportado a '{file_path}' en formato PDF.")
    except Exception as e:
        print(f"Ocurrió un error al exportar a PDF: {e}")

# Función para exportar a Excel
def exportar_a_excel(detalles_pagos, file_path, color1='DDDDDD', color2='FFFFFF'):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Pago", "Cuota", "Saldo", "Interes", "Amortizacion"])
        for i, (pago, cuota, saldo, interes, amortizacion) in enumerate(detalles_pagos):
            color = color1 if i % 2 == 0 else color2
            sheet.append([pago, cuota, saldo, interes, amortizacion])
            for col in range(1, 6): # Columnas de A a E
                cell = sheet.cell(row=i+2, column=col) # +2 para ajustar el índice de fila
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        workbook.save(file_path)
        print(f"Los detalles de cada pago se han exportado a '{file_path}' en formato Excel.")
    except Exception as e:
        print(f"Ocurrió un error al exportar a Excel: {e}")

# Función principal para calcular la amortización de un préstamo
def calcular_amortizacion(factor_amortizacion, tasa_de_interes, valor_presente, numero_pagos):
    detalles_pagos = []  # Lista para almacenar los detalles de cada pago
    saldo = valor_presente  # Saldo inicial del préstamo
    
    # Iterar sobre cada pago
    for pago in range(1, numero_pagos + 1):
        # Calcular el interés del pago
        interes = saldo * tasa_de_interes / 12
        # Calcular la amortización del pago
        amortizacion = factor_amortizacion * valor_presente
        # Calcular la cuota del pago (interés + amortización)
        cuota = interes + amortizacion
        # Actualizar el saldo restante
        saldo -= amortizacion
        # Agregar los detalles del pago a la lista
        detalles_pagos.append((pago, cuota, saldo, interes, amortizacion))
    
    # Calcular el interés total pagado durante el período de amortización
    interes_total_pagado = valor_presente * tasa_de_interes * numero_pagos / 12
    
    # Retornar el interés total pagado y los detalles de cada pago
    return interes_total_pagado, detalles_pagos


# Función principal
def main():
    try:
        # Solicitar al usuario el formato de exportación deseado
        formato_exportacion = input("Seleccione el formato de exportación (CSV/PDF/Excel): ").lower()

        # Especificar la ubicación completa del archivo
        file_name = input("Ingrese el nombre del archivo:") 
        if not file_name.endswith('.' + formato_exportacion):
            file_name += '.' + formato_exportacion
            
        # Calcular el factor de amortización


        # Validar los datos de entrada
        if tasa_de_interes <= 0 or tasa_de_interes >= 1.1:
            raise ValueError("La tasa de interés debe estar entre 0 y 1.")
        if valor_presente <= 0:
            raise ValueError("El valor presente debe ser mayor a cero")
        if numero_pagos <= 0 or numero_pagos > 12:
            raise ValueError("El número de pagos debe ser mayor que 0 y menor o igual que 12")

        # Calcular el factor de amortización
        tasa_interes_periodica = tasa_de_interes / 12
        factor_amortizacion = (1 - (1 + tasa_interes_periodica)**-numero_pagos) / (tasa_interes_periodica * (1 + tasa_interes_periodica)**-numero_pagos)

        # Calcular la amortización y obtener los detalles de cada pago
        interes_total_pagado, detalles_pagos = calcular_amortizacion(factor_amortizacion, tasa_de_interes, valor_presente, numero_pagos)

        # Solicitar al usuario el tipo de gráfico deseado
        tipo_grafico = input("Seleccione el tipo de gráfico (barras_agrupadas/broken_barh/coherencia/fill_between/csd): ").lower()

        # Llamar a la función para seleccionar y mostrar el gráfico
        seleccionar_grafico(tipo_grafico, detalles_pagos)

        # Imprimir el interés total pagado
        print("Interés total pagado:", interes_total_pagado)

        # Mostrar los detalles de cada pago en la consola
        print("Detalles de cada pago:")
        print("Pago\tCuota\tSaldo\tInteres\tAmortizacion")
        for pago, cuota, saldo, interes, amortizacion in detalles_pagos:
            print(f"{pago}\t{cuota:.2f}\t{saldo:.2f}\t{interes:.2f}\t{amortizacion:.2f}")

        if formato_exportacion == 'csv':
            lista_datos = crear_lista_datos(detalles_pagos)
            with open(file_name, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerows(lista_datos)
            print(f"Los detalles de cada pago se han exportado a '{file_name}' en formato CSV.")
        elif formato_exportacion == 'pdf':
            exportar_a_pdf(detalles_pagos, file_name, color1=colors.beige, color2=colors.white)
        elif formato_exportacion == 'excel':
            exportar_a_excel(detalles_pagos, file_name, color1='DDDDDD', color2='FFFFFF')
        else:
            raise ValueError("Formato de exportación no válido. Por favor, seleccione CSV, PDF o Excel.")

    except ValueError as e:
        print("Error:", e)

# Llamar a la función principal
if __name__ == "__main__":
    main()
